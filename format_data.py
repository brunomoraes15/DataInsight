import re
import os
import json
import openpyxl
from openpyxl.utils import get_column_letter
from config import *

def salvar_dados_json(dados: dict, dir_saida: str):
    arquivo_json = f'{conteudo_projeto["nome"]}.json'
    try:
        with open(os.path.join(dir_saida, arquivo_json), 'r', encoding='utf8') as arquivo:
            dados_armazenados = json.load(arquivo)
            
    except FileNotFoundError:
        dados_armazenados = []

    dados_armazenados.append(dados)

    with open(os.path.join(dir_saida, arquivo_json), 'w', encoding='utf8') as arquivo:
        json.dump(dados_armazenados, arquivo, indent=4, ensure_ascii=False)
        print(f'json salvo em {dir_saida+arquivo_json}')

def lattes_formatar_conteudo(conteudo_projeto: dict):
    dados_classificados = {}
    dados_projetos = []
    
    for tipo in conteudo_projeto:
        if tipo in ["pesq", "ext", "dev"]:
            if conteudo_projeto[tipo]: 
                for item in conteudo_projeto[tipo]:
                    texto_formatado = re.split(r'\n', item)
                    dados_projetos.extend(texto_formatado)
                

    dados_projetos = [
        item.strip() for item in dados_projetos
            if item.strip() and not any(item.startswith(palavra) for palavra in palavras_chave_limpeza)
    ]

    dados_projetos_final = []
    for item in dados_projetos:
        dados_projetos_final.extend(item.split(';'))

    dados_projetos_final = [item.strip() for item in dados_projetos_final if item.strip()]

                       
    for info in dados_projetos_final:
        match True:  
            case _ if any(info.startswith(str(ano)) for ano in range(1970, 2025)):
                if dados_classificados:
                    salvar_dados_json(dados_classificados, dir_json)
                    for chave in dados_classificados:
                        dados_classificados[chave] = " "
                
                partes_ano = info.split(' - ')
                dados_classificados['ano_inicio'] = partes_ano[0].strip()
                dados_classificados['ano_fim'] = partes_ano[1].strip()

            case _ if  info.startswith("Situação:"):
                dados_classificados["situacao"] = info.replace('Situação: ', '')

            case _ if info.startswith("Natureza:"):
                dados_classificados["natureza"] = info.replace('Natureza: ', '')

            case _ if info.startswith("Alunos envolvidos:"):
                dados_classificados["alunos"] = info.replace('Alunos envolvidos: ', '')

            case _ if info.startswith("Integrantes:"):
                dados_classificados["integrantes"] = info.replace('Integrantes: ', '')

            case _ if info.startswith("Número de produções C, T & A:"):
                dados_classificados["num_producoes"] = info.replace('Número de produções C, T & A:', '')

            case _ if info.startswith("Financiador(es):"):
                               
                partes_fin = info.split(' - ')
                dados_classificados['financiador'] = partes_fin[0].strip()
                tipo = partes_fin[1].split('.')
                dados_classificados['tipo_financiador'] = tipo[0]
                
                if tipo[1].startswith('Número de orientações: '):
                    dados_classificados['num_orientacoes'] = tipo[1].replace('Número de orientações: ', '')

            case _:
                dados_classificados["titulo"] = info
                dados_classificados['pesquisador'] = conteudo_projeto['nome']


def converter_xlsm(dir_json: str, dir_saida: str):
    for nome_arquivo in os.listdir(dir_json):
        if nome_arquivo.endswith('.json'):
            caminho_arquivo = os.path.join(dir_json, nome_arquivo)
            
            with open(caminho_arquivo, 'r', encoding='utf-8') as arquivo_json:
                dados = json.load(arquivo_json)
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Dados"
                
                for j, coluna in enumerate(dados_classificados.keys()):
                    col_letter = get_column_letter(j + 1)
                    ws[f"{col_letter}1"] = coluna
                
                if isinstance(dados, list):
                    for i, item in enumerate(dados):
                        for j, coluna in enumerate(dados_classificados.keys()):
                            col_letter = get_column_letter(j + 1)
                            ws[f"{col_letter}{i + 2}"] = item.get(coluna, "")
                elif isinstance(dados, dict):
                    for j, coluna in enumerate(dados_classificados.keys()):
                        col_letter = get_column_letter(j + 1)
                        ws[f"{col_letter}2"] = dados.get(coluna, "")
                
                caminho_arquivo_xlsm = os.path.join(dir_saida, f"{os.path.splitext(nome_arquivo)[0]}.xlsm")
                wb.save(caminho_arquivo_xlsm)
                print(f"Planilha XLSM salva em: {caminho_arquivo_xlsm}")
